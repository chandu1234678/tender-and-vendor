"""Excel report builder.

Produces two artefacts:
1. vendor_comparison_matrix.xlsx  – summary Matrix / Details / Summary sheets
2. vendor_<id>.xlsx               – one file per vendor that is an EXACT copy of
   Tech_Comp_check_list.xlsx with the AI compliance results written into the
   Bidder's Compliance (Y/N) / Remarks / Page No. columns, preserving every
   merged cell, style, header row and layout from the original template.
"""
from __future__ import annotations

import copy
import os
import re
import sqlite3
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _status_to_yn(status: Optional[str]) -> str:
    if not status:
        return ""
    s = str(status).strip().upper()
    if s.startswith("YES"):
        return "Y"
    if s.startswith("NO"):
        return "N"
    if s.startswith("NEARLY"):
        return "NEARLY OK"
    return str(status).strip()


def _safe_set(ws, row: int, col: int, value) -> None:
    """Write value to a cell, resolving merged-cell anchors automatically."""
    if row is None or col is None:
        return
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # find the top-left anchor of the merged range that owns this cell
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col, value=value)
                return
        return  # can't resolve – skip
    cell.value = value


def _safe_file_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]", "_", name.strip())
    return cleaned or "file"


def _vendor_output_name(vendor_id: str, history_tag: Optional[str]) -> str:
    safe_vendor = _safe_file_name(vendor_id)
    if history_tag:
        return f"vendor_{safe_vendor}_{history_tag}.xlsx"
    return f"vendor_{safe_vendor}.xlsx"


# ---------------------------------------------------------------------------
# Template-aware per-vendor writer
# ---------------------------------------------------------------------------

def _normalize_vendor_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _vendor_name_matches_id(vendor_name: str, vendor_id: str) -> bool:
    normalized_name = _normalize_vendor_name(vendor_name)
    normalized_id = _normalize_vendor_name(vendor_id)
    if not normalized_name or not normalized_id:
        return False
    if normalized_name == normalized_id:
        return True
    if normalized_name in normalized_id or normalized_id in normalized_name:
        return True
    name_tokens = re.findall(r"[a-z0-9]+", normalized_name)
    id_tokens = re.findall(r"[a-z0-9]+", normalized_id)
    if all(token in normalized_id for token in name_tokens) or all(token in normalized_name for token in id_tokens):
        return True
    return False


def _find_compliance_cols(ws, header_row: int) -> List[Dict[str, Any]]:
    """Return list of vendor column groups found in the header row.

    Each group: {vendor_name, compliance_col, remarks_col, page_col}
    The first group (col 4 in single-vendor sheets) is the primary one.
    Multi-vendor sheets (Mon1 style) have repeating triplets from col 4 onward.
    """
    groups: List[Dict[str, Any]] = []
    max_col = ws.max_column

    def is_compliance_header(text: str) -> bool:
        low = text.lower()
        return (
            ("compliance" in low or "vendor response" in low or "bidder" in low or "response" in low)
            and "remark" not in low
            and "page" not in low
        )

    # Read vendor names from row above the header
    vendor_row = header_row - 1 if header_row > 1 else None

    col = 1
    while col <= max_col:
        val = ws.cell(row=header_row, column=col).value
        if val and isinstance(val, str) and is_compliance_header(val):
            vendor_name = ""
            if vendor_row:
                for lookup_row in range(vendor_row, max(0, vendor_row - 3), -1):
                    v = ws.cell(row=lookup_row, column=col).value
                    if v and isinstance(v, str):
                        low = v.lower()
                        if not any(t in low for t in ["compliance", "remark", "page", "item", "s. no", "bidder"]):
                            vendor_name = v.strip()
                            break
            if not vendor_name:
                vendor_name = f"Vendor_{len(groups) + 1}"
            groups.append({
                "vendor_name": vendor_name,
                "compliance_col": col,
                "remarks_col": col + 1 if col + 1 <= max_col else col,
                "page_col": col + 2 if col + 2 <= max_col else col,
            })
            col += 3
        else:
            col += 1

    if not groups:
        groups.append({
            "vendor_name": "",
            "compliance_col": 4,
            "remarks_col": 5,
            "page_col": 6,
        })
    return groups


def _find_header_row(ws) -> int:
    """Return the 1-based row index of the S.No. / Parameters header."""
    for r in range(1, min(10, ws.max_row) + 1):
        for c in range(1, min(5, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "s. no" in v.lower():
                return r
    return 3  # default


def _build_row_lookup(ws, header_row: int) -> Dict[int, int]:
    """Map serial-number (S.No.) integer → actual Excel row number."""
    lookup: Dict[int, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        try:
            sno = int(v)
            lookup[sno] = r
        except (TypeError, ValueError):
            pass
    return lookup


def _write_vendor_into_sheet(
    ws,
    header_row: int,
    compliance_col: int,
    remarks_col: int,
    page_col: int,
    results_by_sno: Dict[int, Dict[str, Any]],
) -> None:
    """Write AI results into the correct rows of a single sheet."""
    row_lookup = _build_row_lookup(ws, header_row)

    for sno, result in results_by_sno.items():
        excel_row = row_lookup.get(sno)
        if excel_row is None:
            continue
        yn = _status_to_yn(result.get("status"))
        reasoning = str(result.get("reasoning") or "").strip()
        citation = str(result.get("citation") or "").strip()
        page = result.get("citation_page")

        if citation and reasoning:
            remarks = citation if reasoning.lower() in citation.lower() else f"{citation}; {reasoning}"
        else:
            remarks = citation or reasoning

        if len(remarks) > 500:
            remarks = remarks[:497] + "..."

        _safe_set(ws, excel_row, compliance_col, yn)
        _safe_set(ws, excel_row, remarks_col, remarks)
        _safe_set(ws, excel_row, page_col, page)


def _apply_yn_colors(ws, header_row: int, compliance_col: int) -> None:
    """Color the compliance column cells green/yellow/red."""
    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(r, compliance_col)
        if isinstance(cell, MergedCell):
            continue
        val = str(cell.value or "").strip().upper()
        if val == "Y":
            cell.fill = green
        elif val == "N":
            cell.fill = red
        elif "NEARLY" in val:
            cell.fill = yellow


# ---------------------------------------------------------------------------
# Per-vendor file builder (exact template clone)
# ---------------------------------------------------------------------------

def _find_best_vendor_group(groups: List[Dict[str, Any]], vendor_id: str) -> Dict[str, Any]:
    if not groups:
        raise ValueError("No vendor groups available")
    if len(groups) == 1:
        return groups[0]

    best_match = None
    for group in groups:
        if _vendor_name_matches_id(group.get("vendor_name", ""), vendor_id):
            if best_match is None:
                best_match = group
            elif _normalize_vendor_name(group["vendor_name"]) == _normalize_vendor_name(vendor_id):
                best_match = group
                break
            elif len(group.get("vendor_name", "")) > len(best_match.get("vendor_name", "")):
                best_match = group

    return best_match or groups[0]


def _build_vendor_file(
    template_path: str,
    output_path: str,
    vendor_id: str,
    compliance_data: Dict[str, Dict[int, Dict[str, Any]]],
) -> None:
    """Create one vendor output file that mirrors the template exactly.

    compliance_data: {sheet_name -> {serial_no -> result_dict}}
    """
    wb = load_workbook(template_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        groups = _find_compliance_cols(ws, header_row)

        target_group = _find_best_vendor_group(groups, vendor_id)
        sheet_results = compliance_data.get(sheet_name, {})
        if not sheet_results:
            continue

        _write_vendor_into_sheet(
            ws,
            header_row,
            target_group["compliance_col"],
            target_group["remarks_col"],
            target_group["page_col"],
            sheet_results,
        )
        _apply_yn_colors(ws, header_row, target_group["compliance_col"])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Summary / Matrix workbook (unchanged structure)
# ---------------------------------------------------------------------------

def _style_matrix_wb(wb, sheet_names: List[str]) -> None:
    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    title_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")

    if "Matrix" in sheet_names and "Matrix" in wb.sheetnames:
        ws = wb["Matrix"]
        ws.freeze_panes = "E3"
        header_vals = [str(cell.value or "").lower() for cell in ws[2]]
        vendor_start_col = 2
        if "company_requirement" in header_vals:
            vendor_start_col = header_vals.index("company_requirement") + 2
        for cell in ws[1]:
            cell.fill = title_fill
            cell.font = Font(bold=True)
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=vendor_start_col, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value or "").upper()
                cell.alignment = Alignment(wrap_text=True)
                if val.startswith("YES"):
                    cell.fill = green
                elif val.startswith("NEARLY"):
                    cell.fill = yellow
                elif val.startswith("NO"):
                    cell.fill = red

    for sn in sheet_names:
        if sn not in wb.sheetnames:
            continue
        sheet = wb[sn]
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    cell.font = Font(bold=True)
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0,
            )
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 45)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_excel_report(output_path: str, db_path: str = "data/parsed/app.db", history_tag: Optional[str] = None) -> None:
    """Build the compliance report Excel files.

    Produces:
    - <output_path>                              summary matrix workbook
    - <output_dir>/vendor_<id>[_<history_tag>].xlsx  per-vendor files cloned from template
    """
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM compliance_matrix", conn)
    specs = pd.read_sql_query(
        "SELECT spec_id, sheet_name, parameter_name, company_requirement, row_index FROM master_specs",
        conn,
    )
    specs = specs.drop_duplicates(subset=["spec_id"], keep="last") if not specs.empty else specs
    conn.close()

    if df.empty:
        df = pd.DataFrame(columns=[
            "spec_id", "vendor_id", "status", "citation", "reasoning", "confidence",
            "citation_doc_id", "citation_excerpt", "citation_page", "citation_bbox",
        ])

    out_dir = os.path.dirname(output_path) or "."
    os.makedirs(out_dir, exist_ok=True)

    # ------------------------------------------------------------------ #
    # 1. Summary matrix workbook                                           #
    # ------------------------------------------------------------------ #
    pivot = df.pivot(index="spec_id", columns="vendor_id", values="status") if not df.empty else pd.DataFrame()
    details = df.sort_values(["spec_id", "vendor_id"]) if not df.empty else df

    if not specs.empty:
        matrix = specs.merge(pivot, left_on="spec_id", right_index=True, how="left")
    else:
        matrix = pivot.reset_index()

    if not pivot.empty:
        score = (
            pivot.fillna("")
            .apply(lambda col: col.map(
                lambda v: 2 if str(v).upper().startswith("YES") else 1 if str(v).upper().startswith("NEARLY") else 0
            ))
            .sum()
        )
        score_row: Dict[str, Any] = {"spec_id": "SCORE"}
        for col_name in ["sheet_name", "parameter_name", "company_requirement", "row_index"]:
            if col_name in matrix.columns:
                score_row[col_name] = ""
        score_row.update(score.to_dict())
        matrix = pd.concat([matrix, pd.DataFrame([score_row])], ignore_index=True)

    summary = (
        df.groupby("vendor_id", dropna=False)
        .agg(
            total_specs=("spec_id", "count"),
            yes_count=("status", lambda s: s.str.upper().str.startswith("YES").sum()),
            nearly_ok_count=("status", lambda s: s.str.upper().str.startswith("NEARLY").sum()),
            no_count=("status", lambda s: s.str.upper().str.startswith("NO").sum()),
            avg_confidence=("confidence", "mean"),
        )
        .reset_index()
        if not df.empty
        else pd.DataFrame(columns=["vendor_id", "total_specs", "yes_count", "nearly_ok_count", "no_count", "avg_confidence"])
    )
    summary["score"] = summary["yes_count"] * 2 + summary["nearly_ok_count"]
    summary["score_percent"] = summary.apply(
        lambda row: round((row["score"] / (max(1, row["total_specs"]) * 2)) * 100, 2), axis=1
    )
    summary["recommendation"] = summary["score_percent"].apply(
        lambda v: "RECOMMENDED" if v >= 85 else "SHORTLIST" if v >= 70 else "REVIEW"
    )
    summary = summary.sort_values(["score_percent", "avg_confidence"], ascending=[False, False])

    details_export = details.copy()
    if not details_export.empty and not specs.empty:
        details_export = details_export.merge(specs, on="spec_id", how="left")
        keep_cols = [
            "spec_id", "sheet_name", "parameter_name", "company_requirement",
            "vendor_id", "status", "confidence", "citation",
            "citation_page", "citation_bbox", "reasoning",
        ]
        details_export = details_export[[c for c in keep_cols if c in details_export.columns]]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        matrix.to_excel(writer, sheet_name="Matrix", index=False)
        details_export.to_excel(writer, sheet_name="Details", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

    from openpyxl import load_workbook as _lw
    wb = _lw(output_path)
    if "Matrix" in wb.sheetnames:
        ws = wb["Matrix"]
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(row=1, column=1, value="Vendor Compliance Matrix")
    _style_matrix_wb(wb, ["Matrix", "Details", "Summary"])
    wb.save(output_path)

    # ------------------------------------------------------------------ #
    # 2. Per-vendor files cloned from Tech_Comp_check_list.xlsx template  #
    # ------------------------------------------------------------------ #
    template_path = os.path.join("data", "incoming", "Tech_Comp_check_list.xlsx")
    if not os.path.exists(template_path) or df.empty:
        return  # nothing to do

    # Build lookup: vendor_id -> sheet_name -> serial_no -> result
    # serial_no comes from row_index in master_specs
    vendor_data: Dict[str, Dict[str, Dict[int, Dict[str, Any]]]] = {}

    for _, cm_row in df.iterrows():
        vid = cm_row["vendor_id"]
        sid = cm_row["spec_id"]

        # find the spec row to get sheet_name and row_index
        spec_rows = specs[specs["spec_id"] == sid]
        if spec_rows.empty:
            continue
        spec_row = spec_rows.iloc[0]
        sheet_name = str(spec_row.get("sheet_name") or "").strip()
        row_index = spec_row.get("row_index")
        if not sheet_name or row_index is None:
            continue
        try:
            sno = int(row_index)
        except (TypeError, ValueError):
            continue

        vendor_data.setdefault(vid, {}).setdefault(sheet_name, {})[sno] = {
            "status": cm_row.get("status"),
            "citation": cm_row.get("citation_excerpt") or cm_row.get("citation") or "",
            "reasoning": cm_row.get("reasoning") or "",
            "citation_page": cm_row.get("citation_page"),
        }

    for vendor_id, compliance_data in vendor_data.items():
        vendor_file = os.path.join(out_dir, _vendor_output_name(vendor_id, history_tag))
        _build_vendor_file(template_path, vendor_file, vendor_id, compliance_data)
