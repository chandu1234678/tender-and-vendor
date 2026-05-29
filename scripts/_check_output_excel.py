"""Check the latest output Excel files for data quality issues."""
import openpyxl, glob, os, sys
sys.path.insert(0, ".")

INTERNAL_MARKERS = ("heuristic", "consensus rule", "token overlap", "numeric values do not meet")

def check_file(path, sheet_name="NB01"):
    print(f"\n=== {os.path.basename(path)} — sheet [{sheet_name}] ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet {sheet_name} not found. Available: {wb.sheetnames}")
        return
    ws = wb[sheet_name]
    issues = 0
    print(f"  {'S.No':<5} {'Parameter':<22} {'Y/N':<10} {'Remarks (first 60 chars)':<62} Pg")
    print("  " + "-" * 105)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if len(row) < 4:
            continue
        sno    = str(row[0] or "")
        param  = str(row[1] or "")[:21]
        yn     = str(row[3] or "")
        rem    = str(row[4] or "")
        page   = str(row[5] or "") if len(row) > 5 else ""
        if not yn and not rem:
            continue
        flag = ""
        if any(m in rem.lower() for m in INTERNAL_MARKERS):
            flag = "  <<< INTERNAL DEBUG TEXT"
            issues += 1
        print(f"  {sno:<5} {param:<22} {yn:<10} {rem[:60]:<62} {page}{flag}")
    if issues == 0:
        print(f"\n  [OK] No internal debug text found in remarks.")
    else:
        print(f"\n  [WARN] {issues} rows still have internal debug text in remarks.")

# Check latest per-vendor files
for pattern in [
    "data/output/vendor_product-pdf_*.xlsx",
    "data/output/vendor_4AllAnnexures.xlsx",
    "data/output/vendor_Alstonia-merged.xlsx",
]:
    files = sorted(glob.glob(pattern))
    if files:
        check_file(files[-1], "NB01")

# Also check the matrix summary
matrix = "data/output/vendor_comparison_matrix.xlsx"
if os.path.exists(matrix):
    print(f"\n=== {os.path.basename(matrix)} — Summary sheet ===")
    wb = openpyxl.load_workbook(matrix, data_only=True)
    ws = wb["Summary"]
    for row in ws.iter_rows(values_only=True):
        print(" ", [str(v)[:30] if v is not None else "" for v in row])
