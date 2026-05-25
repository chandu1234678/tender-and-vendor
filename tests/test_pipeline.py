import sqlite3
from pathlib import Path

import fitz
import openpyxl
from openpyxl import Workbook

from src.app import run_pipeline


def _build_master_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Spec_ID", "Parameter_Name", "company_Requirement"])
    ws.append(["SPEC-01", "Max Operating Temp", "Must withstand 600C continuously."])
    wb.save(path)


def _build_vendor_pdf(path: Path) -> None:
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "The vendor complies and must withstand 600C continuously.")
    doc.save(str(path))
    doc.close()


def test_pipeline_creates_db_rows_and_report(tmp_path, monkeypatch):
    project_root = tmp_path
    incoming = project_root / "data" / "incoming"
    parsed = project_root / "data" / "parsed"
    output = project_root / "data" / "output"
    incoming.mkdir(parents=True)
    parsed.mkdir(parents=True)
    output.mkdir(parents=True)

    _build_master_workbook(incoming / "master_spec.xlsx")
    _build_vendor_pdf(incoming / "vendorA.pdf")

    def fake_dispatch(spec, vendor_id, blocks, model_name="llama3"):
        citation = blocks[0]["text"] if blocks else ""
        return {
            "spec_id": spec["Spec_ID"],
            "vendor_id": vendor_id,
            "status": "YES",
            "citation": citation,
            "reasoning": "Mocked for test",
            "confidence": 0.99,
            "citation_page": blocks[0]["page"] if blocks else None,
            "citation_bbox": blocks[0]["bbox"] if blocks else None,
            "top_blocks": blocks[:1],
        }

    monkeypatch.setattr(run_pipeline, "PROJECT_ROOT", project_root)
    monkeypatch.setattr(run_pipeline, "dispatch_spec_vendor", fake_dispatch)

    run_pipeline.main(run_id="test-run")

    db_path = parsed / "app.db"
    report_path = output / "vendor_comparison_matrix.xlsx"
    assert db_path.exists()
    assert report_path.exists()

    conn = sqlite3.connect(str(db_path))
    try:
        matrix_count = conn.execute("SELECT COUNT(*) FROM compliance_matrix").fetchone()[0]
        run_row = conn.execute("SELECT status, progress FROM pipeline_runs WHERE run_id=?", ("test-run",)).fetchone()
    finally:
        conn.close()

    assert matrix_count == 1
    assert run_row[0] == "completed"
    assert float(run_row[1]) == 100.0

    wb = openpyxl.load_workbook(report_path)
    assert wb.sheetnames == ["Matrix", "Details", "Summary"]


def test_pipeline_prefers_tech_checklist_workbook(tmp_path, monkeypatch):
    project_root = tmp_path
    incoming = project_root / "data" / "incoming"
    parsed = project_root / "data" / "parsed"
    output = project_root / "data" / "output"
    incoming.mkdir(parents=True)
    parsed.mkdir(parents=True)
    output.mkdir(parents=True)

    _build_master_workbook(incoming / "master_spec.xlsx")
    _build_vendor_pdf(incoming / "vendorA.pdf")
    _build_master_workbook(incoming / "Tech_Comp_check_list.xlsx")

    selected_sources = []

    def fake_parse_master_excel(path):
        selected_sources.append(Path(path).name)
        return [
            {
                "Spec_ID": "SPEC-01",
                "Parameter_Name": "Max Operating Temp",
                "company_Requirement": "Must withstand 600C continuously.",
                "company_Requirement": "Must withstand 600C continuously.",
            }
        ]

    def fake_dispatch(spec, vendor_id, blocks, model_name="llama3"):
        return {
            "spec_id": spec["Spec_ID"],
            "vendor_id": vendor_id,
            "status": "YES",
            "citation": blocks[0]["text"] if blocks else "",
            "reasoning": "Mocked for test",
            "confidence": 0.99,
            "citation_page": blocks[0]["page"] if blocks else None,
            "citation_bbox": blocks[0]["bbox"] if blocks else None,
            "top_blocks": blocks[:1],
        }

    monkeypatch.setattr(run_pipeline, "PROJECT_ROOT", project_root)
    monkeypatch.setattr(run_pipeline, "parse_master_excel", fake_parse_master_excel)
    monkeypatch.setattr(run_pipeline, "dispatch_spec_vendor", fake_dispatch)

    run_pipeline.main(run_id="prefer-master-test")

    assert selected_sources == ["Tech_Comp_check_list.xlsx"]


def test_pipeline_resumable_and_progress_callback(tmp_path, monkeypatch):
    project_root = tmp_path
    incoming = project_root / "data" / "incoming"
    parsed = project_root / "data" / "parsed"
    output = project_root / "data" / "output"
    incoming.mkdir(parents=True)
    parsed.mkdir(parents=True)
    output.mkdir(parents=True)

    _build_master_workbook(incoming / "master_spec.xlsx")
    _build_vendor_pdf(incoming / "vendorA.pdf")

    call_count = {"dispatch": 0}

    def fake_dispatch(spec, vendor_id, blocks, model_name="llama3"):
        call_count["dispatch"] += 1
        return {
            "spec_id": spec["Spec_ID"],
            "vendor_id": vendor_id,
            "status": "YES",
            "citation": blocks[0]["text"] if blocks else "",
            "reasoning": "Mocked for test",
            "confidence": 0.99,
            "citation_page": blocks[0]["page"] if blocks else None,
            "citation_bbox": blocks[0]["bbox"] if blocks else None,
            "top_blocks": blocks[:1],
        }

    monkeypatch.setattr(run_pipeline, "PROJECT_ROOT", project_root)
    monkeypatch.setattr(run_pipeline, "dispatch_spec_vendor", fake_dispatch)

    db_path = parsed / "app.db"
    run_pipeline.init_db(str(db_path))
    conn = sqlite3.connect(str(db_path))
    try:
        conn.execute(
            "INSERT INTO compliance_matrix (spec_id, vendor_id, status, citation, reasoning, confidence) VALUES (?, ?, ?, ?, ?, ?)",
            ("SPEC-01", "vendorA", "YES", "cached", "cached", 0.9),
        )
        conn.commit()
    finally:
        conn.close()

    progress_events = []

    def progress_cb(progress, message):
        progress_events.append((progress, message))

    run_pipeline.main(run_id="resume-test", progress_cb=progress_cb)

    assert call_count["dispatch"] == 0
    assert any("Skipped" in msg for _, msg in progress_events)
